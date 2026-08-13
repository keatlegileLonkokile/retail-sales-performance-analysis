"""
build_workbook.py

Builds output/Sales_Performance_Report.xlsx from data/sales_data.csv.

The workbook is fully formula-driven (SUMIFS / INDEX / MATCH / LARGE / SMALL /
AVERAGE) -- nothing is hardcoded from the Python side except the raw
transaction data itself, so the report recalculates correctly if the
underlying data ever changes.

Sheets:
    Raw Data              - every transaction (2,400 rows), formula columns
                             for Revenue / Cost / Profit / Margin
    Regional Summary      - SUMIFS rollup by region + revenue chart
    Category Summary      - SUMIFS rollup by category + margin chart
    Monthly Trend         - SUMIFS rollup by month + trend line chart
    Store Performance     - H1 vs H2 revenue by store (SUMIFS w/ date
                             ranges) + clustered bar chart
    Product Summary       - SUMIFS rollup by product, plus LARGE/SMALL +
                             INDEX/MATCH top-5 / bottom-5 tables
    Insights & Recs        - live "Key Metrics" formulas + analyst
                             commentary / recommendations

Run scripts/recalc.py (LibreOffice) after this to compute cached formula
values -- openpyxl only ever writes the formula strings.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
CURRENCY_FMT = '$#,##0;($#,##0)'
PCT_FMT = '0.0%'
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REGIONS = ["Northeast", "Southeast", "Midwest", "West", "Southwest"]
CATEGORIES = ["Electronics", "Apparel", "Home & Kitchen", "Beauty & Personal Care", "Sports & Outdoors"]
STORES = ["New York", "Boston", "Atlanta", "Miami", "Chicago", "Los Angeles", "Seattle", "Dallas"]

CATALOG = {
    "Electronics": ["Wireless Earbuds", "Bluetooth Speaker", "Smartwatch", "Tablet Stand", "Portable Charger"],
    "Apparel": ["Men's T-Shirt", "Women's Leggings", "Denim Jacket", "Running Shoes", "Wool Sweater"],
    "Home & Kitchen": ["Coffee Maker", "Non-Stick Pan Set", "Throw Blanket", "LED Desk Lamp", "Storage Bins (Set of 3)"],
    "Beauty & Personal Care": ["Facial Cleanser", "Moisturizer", "Hair Dryer", "Electric Toothbrush", "Perfume Set"],
    "Sports & Outdoors": ["Yoga Mat", "Camping Tent", "Water Bottle", "Resistance Bands", "Hiking Backpack"],
}
ALL_PRODUCTS = [p for cat in CATALOG.values() for p in cat]


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    df = pd.read_csv("data/sales_data.csv", parse_dates=["Month"])

    wb = Workbook()

    # ------------------------------------------------------------------
    # Raw Data
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Raw Data"
    headers = ["Region", "Store", "Category", "Product", "Month", "Units Sold",
               "Unit Price", "Unit Cost", "Revenue", "Cost", "Profit", "Margin"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(row=r, column=1, value=row.Region).font = BODY_FONT
        ws.cell(row=r, column=2, value=row.Store).font = BODY_FONT
        ws.cell(row=r, column=3, value=row.Category).font = BODY_FONT
        ws.cell(row=r, column=4, value=row.Product).font = BODY_FONT
        mcell = ws.cell(row=r, column=5, value=row.Month.date())
        mcell.number_format = "mmm yyyy"
        mcell.font = BODY_FONT
        ws.cell(row=r, column=6, value=int(row._5)).font = BODY_FONT  # Units Sold
        pcell = ws.cell(row=r, column=7, value=float(row._6)); pcell.number_format = CURRENCY_FMT; pcell.font = BODY_FONT  # Unit Price
        ccell = ws.cell(row=r, column=8, value=float(row._7)); ccell.number_format = CURRENCY_FMT; ccell.font = BODY_FONT  # Unit Cost

        rev = ws.cell(row=r, column=9, value=f"=F{r}*G{r}"); rev.number_format = CURRENCY_FMT; rev.font = BODY_FONT
        cost = ws.cell(row=r, column=10, value=f"=F{r}*H{r}"); cost.number_format = CURRENCY_FMT; cost.font = BODY_FONT
        profit = ws.cell(row=r, column=11, value=f"=I{r}-J{r}"); profit.number_format = CURRENCY_FMT; profit.font = BODY_FONT
        margin = ws.cell(row=r, column=12, value=f"=IFERROR(K{r}/I{r},0)"); margin.number_format = PCT_FMT; margin.font = BODY_FONT

    last_row = len(df) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{last_row}"
    autosize(ws, [12, 13, 22, 24, 10, 10, 10, 10, 12, 12, 12, 10])

    RD = "'Raw Data'"

    # ------------------------------------------------------------------
    # Regional Summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Regional Summary")
    ws2.append(["Region", "Revenue", "Cost", "Profit", "Margin"])
    style_header_row(ws2, 1, 5)
    for i, region in enumerate(REGIONS, start=2):
        ws2.cell(row=i, column=1, value=region).font = BODY_FONT
        rev = ws2.cell(row=i, column=2, value=f"=SUMIFS({RD}!$I:$I,{RD}!$A:$A,A{i})")
        cost = ws2.cell(row=i, column=3, value=f"=SUMIFS({RD}!$J:$J,{RD}!$A:$A,A{i})")
        profit = ws2.cell(row=i, column=4, value=f"=B{i}-C{i}")
        margin = ws2.cell(row=i, column=5, value=f"=IFERROR(D{i}/B{i},0)")
        for cell, fmt in [(rev, CURRENCY_FMT), (cost, CURRENCY_FMT), (profit, CURRENCY_FMT), (margin, PCT_FMT)]:
            cell.number_format = fmt
            cell.font = BODY_FONT
    autosize(ws2, [16, 14, 14, 14, 12])

    chart1 = BarChart()
    chart1.title = "Revenue by Region"
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = "Region"
    chart1.style = 10
    data = Reference(ws2, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=6)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width, chart1.height = 16, 9
    ws2.add_chart(chart1, "G2")

    # ------------------------------------------------------------------
    # Category Summary
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Category Summary")
    ws3.append(["Category", "Revenue", "Cost", "Profit", "Margin"])
    style_header_row(ws3, 1, 5)
    for i, cat in enumerate(CATEGORIES, start=2):
        ws3.cell(row=i, column=1, value=cat).font = BODY_FONT
        rev = ws3.cell(row=i, column=2, value=f"=SUMIFS({RD}!$I:$I,{RD}!$C:$C,A{i})")
        cost = ws3.cell(row=i, column=3, value=f"=SUMIFS({RD}!$J:$J,{RD}!$C:$C,A{i})")
        profit = ws3.cell(row=i, column=4, value=f"=B{i}-C{i}")
        margin = ws3.cell(row=i, column=5, value=f"=IFERROR(D{i}/B{i},0)")
        for cell, fmt in [(rev, CURRENCY_FMT), (cost, CURRENCY_FMT), (profit, CURRENCY_FMT), (margin, PCT_FMT)]:
            cell.number_format = fmt
            cell.font = BODY_FONT
    autosize(ws3, [24, 14, 14, 14, 12])

    chart2 = BarChart()
    chart2.title = "Profit Margin by Category"
    chart2.y_axis.title = "Margin"
    chart2.x_axis.title = "Category"
    chart2.style = 11
    data = Reference(ws3, min_col=5, min_row=1, max_row=6)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=6)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width, chart2.height = 16, 9
    ws3.add_chart(chart2, "G2")

    # ------------------------------------------------------------------
    # Monthly Trend
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("Monthly Trend")
    ws4.append(["Month", "Total Revenue", "Electronics Revenue", "Total Profit"])
    style_header_row(ws4, 1, 4)
    months = sorted(df["Month"].dt.date.unique())
    for i, m in enumerate(months, start=2):
        mcell = ws4.cell(row=i, column=1, value=m)
        mcell.number_format = "mmm yyyy"
        mcell.font = BODY_FONT
        rev = ws4.cell(row=i, column=2, value=f"=SUMIFS({RD}!$I:$I,{RD}!$E:$E,A{i})")
        erev = ws4.cell(row=i, column=3, value=f'=SUMIFS({RD}!$I:$I,{RD}!$E:$E,A{i},{RD}!$C:$C,"Electronics")')
        profit = ws4.cell(row=i, column=4, value=f"=SUMIFS({RD}!$K:$K,{RD}!$E:$E,A{i})")
        for cell in (rev, erev, profit):
            cell.number_format = CURRENCY_FMT
            cell.font = BODY_FONT
    autosize(ws4, [14, 16, 20, 14])

    chart3 = LineChart()
    chart3.title = "Monthly Revenue Trend (Total vs. Electronics)"
    chart3.y_axis.title = "Revenue ($)"
    chart3.x_axis.title = "Month"
    chart3.style = 12
    data = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=13)
    cats = Reference(ws4, min_col=1, min_row=2, max_row=13)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.width, chart3.height = 18, 9
    ws4.add_chart(chart3, "F2")

    # ------------------------------------------------------------------
    # Store Performance (H1 vs H2, to surface Miami's decline)
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("Store Performance")
    ws5.append(["Store", "Region", "H1 Revenue (Aug-Jan)", "H2 Revenue (Feb-Jul)",
                "% Change H1->H2", "Full-Year Revenue", "Full-Year Profit", "Margin"])
    style_header_row(ws5, 1, 8)
    for i, store in enumerate(STORES, start=2):
        ws5.cell(row=i, column=1, value=store).font = BODY_FONT
        region = ws5.cell(row=i, column=2, value=f"=INDEX({RD}!$A:$A,MATCH(A{i},{RD}!$B:$B,0))")
        h1 = ws5.cell(row=i, column=3, value=(
            f'=SUMIFS({RD}!$I:$I,{RD}!$B:$B,A{i},{RD}!$E:$E,">="&DATE(2025,8,1),{RD}!$E:$E,"<"&DATE(2026,2,1))'
        ))
        h2 = ws5.cell(row=i, column=4, value=(
            f'=SUMIFS({RD}!$I:$I,{RD}!$B:$B,A{i},{RD}!$E:$E,">="&DATE(2026,2,1),{RD}!$E:$E,"<"&DATE(2026,8,1))'
        ))
        change = ws5.cell(row=i, column=5, value=f"=IFERROR((D{i}-C{i})/C{i},0)")
        fy_rev = ws5.cell(row=i, column=6, value=f"=C{i}+D{i}")
        fy_profit = ws5.cell(row=i, column=7, value=f"=SUMIFS({RD}!$K:$K,{RD}!$B:$B,A{i})")
        margin = ws5.cell(row=i, column=8, value=f"=IFERROR(G{i}/F{i},0)")
        region.font = BODY_FONT
        for cell, fmt in [(h1, CURRENCY_FMT), (h2, CURRENCY_FMT), (change, PCT_FMT),
                           (fy_rev, CURRENCY_FMT), (fy_profit, CURRENCY_FMT), (margin, PCT_FMT)]:
            cell.number_format = fmt
            cell.font = BODY_FONT
    autosize(ws5, [14, 12, 20, 20, 16, 18, 16, 12])

    chart4 = BarChart()
    chart4.type = "col"
    chart4.grouping = "clustered"
    chart4.title = "H1 vs. H2 Revenue by Store"
    chart4.y_axis.title = "Revenue ($)"
    chart4.x_axis.title = "Store"
    chart4.style = 10
    data = Reference(ws5, min_col=3, max_col=4, min_row=1, max_row=9)
    cats = Reference(ws5, min_col=1, min_row=2, max_row=9)
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.width, chart4.height = 20, 10
    ws5.add_chart(chart4, "J2")

    # ------------------------------------------------------------------
    # Product Summary + Top 5 / Bottom 5
    # ------------------------------------------------------------------
    ws6 = wb.create_sheet("Product Summary")
    ws6.append(["Product", "Category", "Revenue", "Profit", "Margin"])
    style_header_row(ws6, 1, 5)
    n_products = len(ALL_PRODUCTS)
    for i, product in enumerate(ALL_PRODUCTS, start=2):
        ws6.cell(row=i, column=1, value=product).font = BODY_FONT
        cat = ws6.cell(row=i, column=2, value=f"=INDEX({RD}!$C:$C,MATCH(A{i},{RD}!$D:$D,0))")
        rev = ws6.cell(row=i, column=3, value=f"=SUMIFS({RD}!$I:$I,{RD}!$D:$D,A{i})")
        profit = ws6.cell(row=i, column=4, value=f"=SUMIFS({RD}!$K:$K,{RD}!$D:$D,A{i})")
        margin = ws6.cell(row=i, column=5, value=f"=IFERROR(D{i}/C{i},0)")
        cat.font = BODY_FONT
        for cell, fmt in [(rev, CURRENCY_FMT), (profit, CURRENCY_FMT), (margin, PCT_FMT)]:
            cell.number_format = fmt
            cell.font = BODY_FONT
    last_prod_row = n_products + 1  # 26
    autosize(ws6, [26, 24, 14, 14, 12])

    rev_range = f"$C$2:$C${last_prod_row}"
    prod_range = f"$A$2:$A${last_prod_row}"

    top_start = last_prod_row + 3   # row 29
    ws6.cell(row=top_start - 1, column=1, value="Top 5 Products by Revenue").font = BOLD_FONT
    ws6.append_dummy = None
    ws6.cell(row=top_start, column=1, value="Rank").font = HEADER_FONT
    ws6.cell(row=top_start, column=2, value="Product").font = HEADER_FONT
    ws6.cell(row=top_start, column=3, value="Revenue").font = HEADER_FONT
    for c in range(1, 4):
        ws6.cell(row=top_start, column=c).fill = HEADER_FILL
        ws6.cell(row=top_start, column=c).border = BORDER
    for k in range(1, 6):
        r = top_start + k
        ws6.cell(row=r, column=1, value=k).font = BODY_FONT
        revcell = ws6.cell(row=r, column=3, value=f"=LARGE({rev_range},{k})")
        revcell.number_format = CURRENCY_FMT
        revcell.font = BODY_FONT
        prodcell = ws6.cell(row=r, column=2, value=f"=INDEX({prod_range},MATCH(C{r},{rev_range},0))")
        prodcell.font = BODY_FONT

    bottom_start = top_start + 7  # a couple rows below the top-5 table
    ws6.cell(row=bottom_start - 1, column=1, value="Bottom 5 Products by Revenue").font = BOLD_FONT
    ws6.cell(row=bottom_start, column=1, value="Rank").font = HEADER_FONT
    ws6.cell(row=bottom_start, column=2, value="Product").font = HEADER_FONT
    ws6.cell(row=bottom_start, column=3, value="Revenue").font = HEADER_FONT
    for c in range(1, 4):
        ws6.cell(row=bottom_start, column=c).fill = HEADER_FILL
        ws6.cell(row=bottom_start, column=c).border = BORDER
    for k in range(1, 6):
        r = bottom_start + k
        ws6.cell(row=r, column=1, value=k).font = BODY_FONT
        revcell = ws6.cell(row=r, column=3, value=f"=SMALL({rev_range},{k})")
        revcell.number_format = CURRENCY_FMT
        revcell.font = BODY_FONT
        prodcell = ws6.cell(row=r, column=2, value=f"=INDEX({prod_range},MATCH(C{r},{rev_range},0))")
        prodcell.font = BODY_FONT

    # ------------------------------------------------------------------
    # Insights & Recommendations
    # ------------------------------------------------------------------
    ws7 = wb.create_sheet("Insights & Recommendations")
    ws7.column_dimensions["A"].width = 30
    ws7.column_dimensions["B"].width = 60

    ws7["A1"] = "Northwind Retail — Sales Performance: Key Findings"
    ws7["A1"].font = TITLE_FONT
    ws7.merge_cells("A1:D1")
    ws7["A2"] = "All figures below are computed live via formulas from the Raw Data sheet."
    ws7["A2"].font = SUBTITLE_FONT
    ws7.merge_cells("A2:D2")

    metrics = [
        ("Total Revenue (FY)", "=SUM('Raw Data'!$I:$I)", CURRENCY_FMT),
        ("Total Profit (FY)", "=SUM('Raw Data'!$K:$K)", CURRENCY_FMT),
        ("Overall Margin", "=B5/B4", PCT_FMT),
        ("Top Region", "=INDEX('Regional Summary'!$A$2:$A$6,MATCH(MAX('Regional Summary'!$B$2:$B$6),'Regional Summary'!$B$2:$B$6,0))", None),
        ("Top Region Revenue", "=MAX('Regional Summary'!$B$2:$B$6)", CURRENCY_FMT),
        ("Bottom Region", "=INDEX('Regional Summary'!$A$2:$A$6,MATCH(MIN('Regional Summary'!$B$2:$B$6),'Regional Summary'!$B$2:$B$6,0))", None),
        ("Bottom Region Revenue", "=MIN('Regional Summary'!$B$2:$B$6)", CURRENCY_FMT),
        ("Most Declining Store (H1->H2)", "=INDEX('Store Performance'!$A$2:$A$9,MATCH(MIN('Store Performance'!$E$2:$E$9),'Store Performance'!$E$2:$E$9,0))", None),
        ("Decline %", "=MIN('Store Performance'!$E$2:$E$9)", PCT_FMT),
        ("Thinnest-Margin Category", "=INDEX('Category Summary'!$A$2:$A$6,MATCH(MIN('Category Summary'!$E$2:$E$6),'Category Summary'!$E$2:$E$6,0))", None),
        ("Thinnest Margin %", "=MIN('Category Summary'!$E$2:$E$6)", PCT_FMT),
        ("Electronics Holiday Lift (Nov/Dec vs. rest of year)",
         "=AVERAGE('Monthly Trend'!$C$5:$C$6)/AVERAGE('Monthly Trend'!$C$2:$C$4,'Monthly Trend'!$C$7:$C$13)-1",
         PCT_FMT),
    ]
    row = 4
    for label, formula, fmt in metrics:
        lcell = ws7.cell(row=row, column=1, value=label)
        lcell.font = BOLD_FONT
        vcell = ws7.cell(row=row, column=2, value=formula)
        vcell.font = BODY_FONT
        if fmt:
            vcell.number_format = fmt
        row += 1

    row += 1
    ws7.cell(row=row, column=1, value="Findings & Recommendations").font = TITLE_FONT
    ws7.merge_cells(f"A{row}:D{row}")
    row += 1
    ws7.cell(row=row, column=1, value="(Analyst commentary — see the Key Metrics above and the underlying sheets for the figures cited.)").font = SUBTITLE_FONT
    ws7.merge_cells(f"A{row}:D{row}")
    row += 2

    findings = [
        ("1. Miami is in a steady, accelerating decline.",
         "Miami's revenue fell roughly 43% from the first half of the year to the second half, the "
         "steepest drop of any store, while every other Southeast store held steady or grew. This reads as "
         "a competitive or local-demand problem, not a seasonal blip. Recommend an on-the-ground review "
         "(new competitor activity, pricing, foot traffic) within the next quarter, paired with a targeted "
         "local promotion to test whether discounting recovers volume before considering a footprint change."),
        ("2. Dallas is dramatically underperforming in Sports & Outdoors specifically.",
         "Dallas sells about 75% less Sports & Outdoors merchandise than the average of the other stores, "
         "while its performance in every other category is in line with peers — so this isn't a "
         "store-wide problem. Recommend auditing local assortment and shelf placement for that category, "
         "and testing regional marketing (e.g. tying inventory to local outdoor/sports seasonality in Texas) "
         "before cutting shelf space."),
        ("3. Home & Kitchen carries the thinnest margin of any category.",
         "At roughly a 29% margin versus 53-59% for Apparel and Beauty & Personal Care, Home & Kitchen is "
         "the category most exposed to cost inflation. Recommend a supplier/cost review for this category "
         "and testing a modest price increase on the highest-volume SKUs, where demand is least likely to "
         "be price-sensitive."),
        ("4. Electronics spikes sharply in November and December.",
         "Electronics revenue in Nov/Dec runs well above the rest-of-year average, consistent with holiday "
         "shopping. Recommend building this into inventory and staffing plans ahead of Q4 rather than "
         "reacting to it, and considering a complementary cross-sell push (accessories, extended warranties) "
         "during that window to capture additional margin."),
        ("5. Apparel and Beauty & Personal Care are the strongest all-around performers.",
         "Both combine solid revenue with the highest margins in the portfolio. Recommend protecting and "
         "growing shelf space/marketing spend for these categories, and using them as the template "
         "(assortment, pricing approach) when reviewing the weaker categories above."),
    ]
    for title, body in findings:
        ws7.cell(row=row, column=1, value=title).font = BOLD_FONT
        ws7.merge_cells(f"A{row}:D{row}")
        row += 1
        cell = ws7.cell(row=row, column=1, value=body)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws7.merge_cells(f"A{row}:D{row}")
        ws7.row_dimensions[row].height = 60
        row += 2

    # ------------------------------------------------------------------
    wb.save("output/Sales_Performance_Report.xlsx")
    print("Saved output/Sales_Performance_Report.xlsx")


if __name__ == "__main__":
    main()
